from __future__ import annotations

import csv
import io

SUPPORTED_SUFFIXES = (".xlsx", ".xlsm", ".csv", ".pdf")


def is_supported(filename: str) -> bool:
    lowered = (filename or "").lower()
    return lowered.endswith(SUPPORTED_SUFFIXES)


def text_from_csv(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            decoded = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        return ""
    rows = []
    for row in csv.reader(io.StringIO(decoded)):
        rows.append(" ".join(cell for cell in row if cell))
    return "\n".join(rows)


def text_from_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        rows = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell) for cell in row if cell is not None]
                if cells:
                    rows.append(" ".join(cells))
        return "\n".join(rows)
    finally:
        workbook.close()


def text_from_pdf(data: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    pages = []
    for page in reader.pages:
        extracted = page.extract_text() or ""
        if extracted:
            pages.append(extracted)
    return "\n".join(pages)


def extract_text(filename: str, data: bytes) -> str:
    lowered = (filename or "").lower()
    try:
        if lowered.endswith(".csv"):
            return text_from_csv(data)
        if lowered.endswith((".xlsx", ".xlsm")):
            return text_from_xlsx(data)
        if lowered.endswith(".pdf"):
            return text_from_pdf(data)
    except Exception:
        return ""
    return ""
